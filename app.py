import streamlit as st
import pandas as pd
import json
import io
from datetime import datetime, timezone
from sources import fetch_all_sources, build_unified_index
from matcher import match_single, match_batch, normalize_name
from report import (
    generate_single_search_report,
    generate_bulk_screening_report,
)
from config import SOURCE_LABELS, SOURCE_URLS, COLORS

st.set_page_config(
    page_title="SENTINEL | Sanctions Screening Tool",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Session state defaults
_DEFAULTS = {
    "cache_buster": 0,
    "single_result": None,
    "batch_results": None,
    "batch_meta": None,
    "batch_page": 0,
    "search_page": 0,
    "fuzzy_threshold": 85,
    "enable_phonetic": True,
    "last_query": "",
}
for key, val in _DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = val





@st.cache_data(ttl=3600, show_spinner="Loading sanctions lists...")
def load_all_data(cache_buster: int = 0, _version: int = 3):
    data = fetch_all_sources()
    index = build_unified_index(data)
    return data, index


def format_timestamp(ts_iso: str) -> str:
    if not ts_iso:
        return "—"
    try:
        dt = datetime.fromisoformat(ts_iso.replace("Z", "+00:00"))
        now = datetime.now(timezone.utc)
        diff = now - dt
        mins = int(diff.total_seconds() / 60)
        if mins < 1:
            return "just now"
        if mins < 60:
            return f"{mins} min ago"
        hours = int(mins / 60)
        if hours < 24:
            return f"{hours}h ago"
        return dt.strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return ts_iso[:19]


def status_label(status: str) -> str:
    return {"live": "LIVE", "cached": "CACHED", "failed": "FAILED"}.get(status, "UNKNOWN")


def verdict_badge_html(verdict: str) -> str:
    if verdict == "HIT":
        color = "#C83232"
        bg = "#FDECEC"
    elif verdict == "POTENTIAL HIT":
        color = "#E5A913"
        bg = "#FFF8E1"
    else:
        color = "#32B464"
        bg = "#E8F8F0"
    return f'<span style="background:{bg};color:{color};padding:4px 14px;border-radius:12px;font-weight:700;font-size:14px">{verdict}</span>'


st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
        html, body, [class*="css"]  {{
            font-family: 'Poppins', sans-serif;
        }}
        .main .block-container {{ padding-top: 1rem; }}
        .stTabs [data-baseweb="tab"] {{ font-size: 16px; font-weight: 600; }}
        .stTabs [aria-selected="true"] {{ color: #E5A913 !important; }}
        .match-card {{
            background: #F5F5F5; border: 1px solid #333333; border-radius: 8px;
            padding: 12px; margin: 8px 0;
        }}
        .match-card h4 {{ margin: 0 0 6px 0; color: #111111; }}
        .source-card {{
            background: #F5F5F5; border: 1px solid #333333; border-radius: 8px;
            padding: 12px; margin: 4px 0;
        }}
        .stButton > button {{
            background-color: #E5A913; color: #111111; font-weight: 600;
            border: none; border-radius: 6px;
        }}
        .stButton > button:hover {{ background-color: #D09C10; }}
        .stDownloadButton button {{ background-color: #E5A913 !important; color: #111111 !important; font-weight: 600 !important; border: none !important; border-radius: 6px !important; }}
        .stDownloadButton button:hover {{ background-color: #D09C10 !important; }}
        .sidebar-status {{ padding: 2px 0; font-size: 14px; }}
        div[data-testid="stSidebarUserContent"] {{ padding-top: 1rem; }}
        section[data-testid="stSidebar"] {{
            border-right: 4px solid #E5A913;
        }}
    </style>
    """,
    unsafe_allow_html=True,
)


def render_sidebar(data: dict, index: dict):
    with st.sidebar:
        st.markdown("### SENTINEL")
        st.caption("Sanctions screening tool")

        st.divider()
        st.markdown("#### Data Status")

        sources = data.get("sources", {})
        for src_key, src_data in sources.items():
            label = SOURCE_LABELS.get(src_key, src_key)
            status = src_data.get("status", "unknown")
            count = src_data.get("record_count", 0)
            fetched = format_timestamp(src_data.get("fetched_at", ""))
            status_tag = status_label(status)
            display = f"{label}: {count:,} [{status_tag}]"
            if src_key == "mas_investor_alert":
                display += ' <span title="Unlicensed entities flagged by MAS as possibly misrepresenting their status. Not a sanctions list.">ⓘ</span>'
            st.markdown(f'<div class="sidebar-status">{display}</div>', unsafe_allow_html=True)

        st.caption(f"Total: {data.get('total_records', 0):,} records across {len(sources)} lists")
        st.divider()

        if st.button("Refresh Data Now", use_container_width=True):
            st.cache_data.clear()
            st.session_state.cache_buster += 1
            st.session_state.single_result = None
            st.session_state.batch_results = None
            st.rerun()

        with st.expander("Settings"):
            st.session_state.fuzzy_threshold = st.slider(
                "Fuzzy threshold", 75, 100,
                st.session_state.fuzzy_threshold, 1
            )
            st.session_state.enable_phonetic = st.toggle(
                "Phonetic matching", st.session_state.enable_phonetic
            )
            st.session_state.screened_by = st.text_input("Screened by", value="")


def render_degradation_banner(data: dict):
    sources = data.get("sources", {})
    warnings = []
    errors = []
    for src_key, src_data in sources.items():
        label = SOURCE_LABELS.get(src_key, src_key)
        status = src_data.get("status", "")
        if status == "cached":
            warnings.append(f"{label} is using CACHED data (live fetch failed)")
        elif status == "failed":
            errors.append(f"{label} FAILED to load")
    if warnings:
        st.warning(" | ".join(warnings))
    if errors:
        st.error(" | ".join(errors))


def tab_individual_search(index: dict, data: dict):
    records = index.get("records", [])
    name_pairs = index.get("name_pairs", [])

    query = st.text_input(
        "Enter entity name to screen",
        placeholder="Type a name (minimum 2 characters)",
        key="search_input",
    )

    if query and len(query) >= 2:
        norm_query = normalize_name(query)
        suggestions = []
        first_char = norm_query[:3].upper() if norm_query else ""
        if first_char:
            for name, norm in name_pairs[:30000]:
                if first_char in norm[:20]:
                    suggestions.append(name)
                    if len(suggestions) >= 8:
                        break

        col_a, col_b = st.columns([3, 1])
        with col_a:
            if suggestions:
                selected = st.selectbox(
                    "Suggestions",
                    [""] + suggestions,
                    label_visibility="collapsed",
                    key="autocomplete_box",
                )
                if selected:
                    query = selected
                    st.session_state.search_input = selected
                    st.rerun()

        with col_b:
            search_clicked = st.button("Screen", type="primary", use_container_width=True, key="search_btn")

        do_search = search_clicked or (
            len(query.strip()) >= 2
            and st.session_state.get("single_result") is None
        )

        if do_search or st.session_state.get("last_query") != query:
            threshold = st.session_state.fuzzy_threshold
            enable_phonetic = st.session_state.enable_phonetic
            with st.spinner("Screening..."):
                result = match_single(query, records, threshold, enable_phonetic)
            st.session_state.single_result = result
            st.session_state.last_query = query
            st.session_state.search_page = 0

    result = st.session_state.get("single_result")
    if not result or st.session_state.get("last_query") != query:
        return

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(verdict_badge_html(result["verdict"]), unsafe_allow_html=True)

    meta_cols = st.columns(4)
    meta_cols[0].metric("Records Screened", f"{result['total_records_screened']:,}")
    meta_cols[1].metric("Matches", result["match_count"])
    meta_cols[2].metric("Top Confidence", f"{result['top_confidence']}%")
    meta_cols[3].metric("Verdict", result["verdict"])

    if result["matches"]:
        sources = data.get("sources", {})
        pdf_bytes = generate_single_search_report(query, result, sources)
        dl_cols = st.columns([2, 3, 2])
        dl_cols[0].download_button(
            "Download PDF",
            pdf_bytes,
            file_name=f"sanctions_search_{query[:30].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    matches_list = result.get("matches", [])
    sp_size = 50
    sp_total = max(1, (len(matches_list) + sp_size - 1) // sp_size)
    sp_current = min(st.session_state.get("search_page", 0), sp_total - 1)
    sp_current = max(0, sp_current)
    sp_start = sp_current * sp_size
    sp_end = min(sp_start + sp_size, len(matches_list))

    for i, m in enumerate(matches_list[sp_start:sp_end]):
        rec = m["record"]
        conf = m["confidence"]
        strategy = m["strategy"]
        source_label = rec.get("source_label", rec.get("source", ""))
        programs = "; ".join(rec.get("programs", [])[:3])
        aliases = "; ".join(rec.get("aliases", [])[:3])

        with st.expander(f"**{rec.get('name', '')}** — {conf}% — {source_label}", expanded=(i == 0)):
            info_cols = st.columns([1, 1, 1])
            info_cols[0].metric("Confidence", f"{conf}%")
            info_cols[1].metric("Entity Type", rec.get("entity_type", "").title())
            info_cols[2].metric("List", source_label)
            st.markdown(f"**Matched value:** `{m['matched_name']}`")
            if programs:
                st.markdown(f"**Programs:** {programs}")
            if aliases:
                st.markdown(f"**Aliases:** {aliases}")
            if rec.get("remarks"):
                st.markdown(f"**Remarks:** {rec['remarks']}")
            with st.expander("View raw record"):
                st.json({k: v for k, v in rec.items() if not k.startswith("_")})

    if len(matches_list) > sp_size:
        sp_col_prev, sp_col_info, sp_col_next = st.columns([1, 2, 1])
        with sp_col_prev:
            if st.button("Previous", key="sp_prev", disabled=(sp_current == 0), use_container_width=True):
                st.session_state.search_page = sp_current - 1
                st.rerun()
        with sp_col_info:
            st.markdown(f"<div style='text-align:center;padding-top:6px;font-size:14px'>Match {sp_start + 1}–{sp_end} of {len(matches_list)}</div>", unsafe_allow_html=True)
        with sp_col_next:
            if st.button("Next", key="sp_next", disabled=(sp_current >= sp_total - 1), use_container_width=True):
                st.session_state.search_page = sp_current + 1
                st.rerun()

_NAME_KEYWORDS = [
    "name", "full name", "last name", "surname", "client", "customer", "applicant",
    "party", "counterparty", "entity", "individual", "company", "organization",
    "borrower", "supplier", "tenant", "contact", "member",
]


def _norm_header(c) -> str:
    return str(c).strip().lower() if c is not None else ""


_NAME_SPECIFIC = [
    "full name", "last name", "first name", "surname", "client name", "fullname",
]


def _name_col_score(c) -> int:
    h = _norm_header(c)
    if not h:
        return -1
    score = 0
    for kw in _NAME_SPECIFIC:
        if kw in h:
            score += 3
    for kw in _NAME_KEYWORDS:
        if kw in h:
            score += 1
    return score


def _read_csv_bytes(raw: bytes) -> pd.DataFrame:
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    for sep in (",", ";", "\t"):
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep)
            if df.shape[1] > 1:
                return df
        except Exception:
            continue
    return pd.read_csv(io.StringIO(text), sep=",", engine="python")


def _read_excel_workbook(raw: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    out = {}
    for ws in wb.worksheets:
        out[ws.title] = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _read_xls_workbook(raw: bytes) -> dict:
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    out = {}
    for sh in book.sheets():
        out[sh.name] = [sh.row_values(i) for i in range(sh.nrows)]
    return out


def _best_name_column(df: pd.DataFrame):
    best, best_score = None, -1.0
    for c in df.columns:
        vals = df[c].dropna().astype(str).str.strip()
        vals = vals[vals != ""]
        if len(vals) < 5:
            continue
        text_ratio = vals.apply(lambda v: not any(ch.isdigit() for ch in v)).mean()
        avg_len = vals.str.len().mean()
        if text_ratio >= 0.7 and avg_len >= 3:
            score = text_ratio * avg_len
            if score > best_score:
                best, best_score = c, score
    return best


def _clean(text) -> str:
    if text is None:
        return ""
    t = str(text).replace("_x000D_", " ").replace("_x000D", " ")
    t = t.replace("\uFFFD", "")
    return t


def build_match_details_xlsx(results: dict) -> bytes:
    import math

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    detail_cols = [
        ("Client Name", 36),
        ("Verdict", 16),
        ("Source", 30),
        ("Confidence", 11),
        ("Matched Name", 42),
        ("Is Alias", 9),
        ("DOB", 16),
        ("Nationality", 22),
        ("Programs", 45),
        ("Aliases", 60),
        ("Remarks", 90),
    ]
    wrap_cols = {"Programs", "Aliases", "Remarks"}

    data = []
    for r in results.get("results", []):
        query = r.get("query", "")
        verdict = r.get("verdict", "")
        for m in r.get("matches", []):
            rec = m.get("record", {})
            data.append({
                "Client Name": _clean(query),
                "Verdict": verdict,
                "Source": _clean(rec.get("source_label", rec.get("source", ""))),
                "Confidence": m.get("confidence", 0),
                "Matched Name": _clean(m.get("matched_name") or rec.get("name", "")),
                "Is Alias": m.get("is_alias", False),
                "DOB": _clean(rec.get("dob", "")),
                "Nationality": _clean(rec.get("nationality", "")),
                "Programs": _clean("; ".join(rec.get("programs", []))),
                "Aliases": _clean("; ".join(rec.get("aliases", []))),
                "Remarks": _clean(rec.get("remarks", "")),
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Match Details"

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    band_fill = PatternFill("solid", fgColor="F4F4F4")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (name, _) in enumerate(detail_cols, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(data, start=2):
        for ci, (name, _) in enumerate(detail_cols, start=1):
            val = row[name]
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            if name == "Confidence":
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif name in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.alignment = Alignment(vertical="top")
            if ri % 2 == 0:
                c.fill = band_fill

    widths = {name: base for name, base in detail_cols}
    for name, base in detail_cols:
        if name in wrap_cols:
            continue
        maxlen = base
        for row in data:
            maxlen = max(maxlen, len(str(row[name])))
        widths[name] = min(maxlen + 2, 60)
    for ci, (name, _) in enumerate(detail_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[name]

    for ri, row in enumerate(data, start=2):
        lines = 1
        for name in wrap_cols:
            text = str(row[name])
            if text:
                lines = max(lines, math.ceil(len(text) / widths[name]) + text.count("\n"))
        ws.row_dimensions[ri].height = max(16, lines * 14 + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(detail_cols))}{len(data) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def tab_bulk_screening(index: dict, data: dict):
    records = index.get("records", [])

    uploaded_file = st.file_uploader(
        "Upload a CSV, XLSX, or JSON file",
        type=["csv", "xlsx", "xls", "json"],
    )

    df = None
    sheet_rows = {}
    if uploaded_file is not None:
        fname = uploaded_file.name.lower()
        try:
            if fname.endswith(".csv"):
                df = _read_csv_bytes(uploaded_file.getvalue())
            elif fname.endswith(".xlsx"):
                sheet_rows = _read_excel_workbook(uploaded_file.getvalue())
            elif fname.endswith(".xls"):
                sheet_rows = _read_xls_workbook(uploaded_file.getvalue())
            elif fname.endswith(".json"):
                df = pd.DataFrame(json.loads(uploaded_file.read().decode("utf-8")))
        except Exception as e:
            st.error(f"Could not parse file: {e}")
            return

    if sheet_rows:
        sheet_names = list(sheet_rows.keys())
        default_sheet = max(sheet_names, key=lambda s: len(sheet_rows[s]))
        selected_sheet = st.selectbox(
            "Select sheet",
            sheet_names,
            index=sheet_names.index(default_sheet),
        )
        rows = sheet_rows[selected_sheet]
        st.caption(f"Loaded {len(rows)} raw rows from sheet \"{selected_sheet}\"")
        if rows:
            header_row = st.number_input(
                "Header is on row…",
                min_value=1,
                max_value=min(len(rows), 20),
                value=1,
                step=1,
            )
            header = rows[header_row - 1]
            ncols = len(header)
            data_rows = []
            for r in rows[header_row:]:
                r = list(r)
                if len(r) < ncols:
                    r = r + [None] * (ncols - len(r))
                data_rows.append(r[:ncols])
            df = pd.DataFrame(data_rows, columns=header).dropna(how="all")

    if df is not None and not df.empty:
        st.markdown(f"**Preview** ({len(df)} data rows)")
        st.dataframe(df.head(10), use_container_width=True)

        scored = [(c, _name_col_score(c)) for c in df.columns]
        name_cols = [c for c, s in sorted(scored, key=lambda x: x[1], reverse=True) if s > 0]
        if not name_cols:
            fallback = _best_name_column(df)
            if fallback is not None:
                name_cols = [fallback]
        name_col = st.selectbox(
            "Select name column",
            name_cols if name_cols else df.columns,
            index=0,
        )

        id_col = st.selectbox("Select reference/ID column (optional)", ["None"] + list(df.columns))

        if st.button("Run Screening", type="primary", use_container_width=True):
            names = df[name_col].dropna().astype(str).str.strip().tolist()
            names = [n for n in names if len(n) >= 2]

            if not names:
                st.warning("No valid names found in the selected column")
                return

            progress_bar = st.progress(0, text="Screening...")
            status_text = st.empty()

            def on_progress(current, total):
                progress_bar.progress(current / total)
                status_text.text(f"Screening {current}/{total}...")

            threshold = st.session_state.fuzzy_threshold
            enable_phonetic = st.session_state.enable_phonetic
            results = match_batch(names, records, threshold, enable_phonetic, progress_callback=on_progress)

            progress_bar.empty()
            status_text.empty()

            st.session_state.batch_results = results
            st.session_state.batch_page = 0
            st.session_state.batch_meta = {
                "filename": uploaded_file.name,
                "name_column": name_col,
                "id_column": id_col,
            }

    results = st.session_state.get("batch_results")
    if results is None:
        return

    total = results["total"]
    hits = results["hit_count"]
    potentials = results["potential_count"]
    cleans = results["no_hit_count"]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total", total)
    c2.metric("FLAGGED", hits, delta_color="off")
    c3.metric("POTENTIAL", potentials, delta_color="off")
    c4.metric("CLEAR", cleans, delta_color="off")

    show_all = st.checkbox("Show all results", value=False)
    rows = []
    for r in results["results"]:
        if not show_all and r["verdict"] == "NO HIT":
            continue
        rows.append({
            "Name": r["query"],
            "Verdict": r["verdict"],
            "Confidence": r["top_confidence"],
            "Matches": r["match_count"],
        })

    df_results = pd.DataFrame(rows)

    page_size = 50
    total_pages = max(1, (len(df_results) + page_size - 1) // page_size)
    current_page = min(st.session_state.batch_page, total_pages - 1)
    current_page = max(0, current_page)
    start_idx = current_page * page_size
    end_idx = min(start_idx + page_size, len(df_results))
    df_page = df_results.iloc[start_idx:end_idx]

    col_prev, col_info, col_next = st.columns([1, 2, 1])
    with col_prev:
        if st.button("Previous", disabled=(current_page == 0), use_container_width=True):
            st.session_state.batch_page = current_page - 1
            st.rerun()
    with col_info:
        st.markdown(f"<div style='text-align:center;padding-top:6px;font-size:14px'>Showing {start_idx + 1}–{end_idx} of {len(df_results)} (Page {current_page + 1} of {total_pages})</div>", unsafe_allow_html=True)
    with col_next:
        if st.button("Next", disabled=(current_page >= total_pages - 1), use_container_width=True):
            st.session_state.batch_page = current_page + 1
            st.rerun()

    def color_verdict(val):
        if val == "HIT":
            return "color: #C83232; font-weight: bold"
        if val == "POTENTIAL HIT":
            return "color: #E5A913; font-weight: bold"
        return "color: #32B464"

    st.dataframe(
        df_page.style.map(color_verdict, subset=["Verdict"]),
        use_container_width=True,
    )

    meta = st.session_state.get("batch_meta", {})
    sources = data.get("sources", {})
    pdf_bytes = generate_bulk_screening_report(results, sources, meta)

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.download_button(
            "Download Full Report (PDF)",
            pdf_bytes,
            file_name=f"sanctions_bulk_screening_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with col_b:
        csv_buffer = io.StringIO()
        df_results.to_csv(csv_buffer, index=False)
        st.download_button(
            "Download Results (CSV)",
            csv_buffer.getvalue(),
            file_name=f"sanctions_results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with col_c:
        xlsx_bytes = build_match_details_xlsx(results)
        st.download_button(
            "Download Match Details (XLSX)",
            xlsx_bytes,
            file_name=f"sanctions_match_details_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def tab_data_status(data: dict):
    st.markdown("#### Data Source Status")
    st.caption("Last refreshed: " + format_timestamp(data.get("fetch_timestamp", "")))

    sources = data.get("sources", {})
    for src_key, src_data in sources.items():
        label = SOURCE_LABELS.get(src_key, src_key)
        status = src_data.get("status", "unknown")
        count = src_data.get("record_count", 0)
        fetched = src_data.get("fetched_at", "")
        error = src_data.get("error")

        with st.container():
            url = SOURCE_URLS.get(src_key, "")
            link = f'<a href="{url}" target="_blank" style="color:#111;text-decoration:none">{label}</a>' if url else label
            note = ""
            if src_key == "mas_investor_alert":
                note = ' <span style="cursor:help" title="Unlicensed entities flagged by MAS as possibly misrepresenting their status. Not a sanctions list.">ⓘ</span>'
            st.markdown(
                f'<div class="source-card">'
                f'<h4>{link}{note}</h4>'
                f'Status: <b>{status.upper()}</b> | Records: <b>{count:,}</b> | Fetched: <b>{format_timestamp(fetched)}</b>'
                + (f'<br><span style="color:#C83232">Error: {error}</span>' if error else "")
                + '</div>',
                unsafe_allow_html=True,
            )


def tab_about():
    st.markdown("### About SENTINEL")
    st.markdown(
        """
        This tool screens names against **eight international sanctions, enforcement and investor-alert lists**:

        1. **OFAC SDN List** — US Treasury's list of specially designated nationals and blocked persons
        2. **OFAC Non-SDN List** — Consolidated non-SDN sanctions list
        3. **UN Consolidated List** — United Nations Security Council sanctions list
        4. **Australian DFAT List** — Australia's consolidated sanctions list
        5. **MAS Designated Lists** — Singapore's lists of designated individuals and entities
        6. **MAS Enforcement Actions** — Entities subject to MAS enforcement actions
        7. **MAS Investor Alert List** — Unlicensed entities flagged by MAS for potentially misrepresenting their status (investor-protection list, not a sanctions list)
        8. **SG TSFA First Schedule** — Individuals and entities designated under Singapore's Terrorism (Suppression of Financing) Act

        **Matching methodology:** Five-strategy engine — exact, alias, substring, fuzzy (token-set), and phonetic (Double Metaphone).

        **Important:** Results are advisory and require human review. Absence of a match does not confirm a subject is unsanctioned.
        """,
        unsafe_allow_html=True,
    )


def main():
    try:
        data, index = load_all_data(cache_buster=st.session_state.cache_buster)
    except Exception as e:
        st.error(f"Failed to load sanctions data: {e}")
        st.info("The app requires an internet connection to fetch sanctions lists. Please try again.")
        return

    render_sidebar(data, index)
    render_degradation_banner(data)

    tab1, tab2, tab3, tab4 = st.tabs(["Individual Search", "Bulk Screening", "Data Status", "About"])

    with tab1:
        tab_individual_search(index, data)
    with tab2:
        tab_bulk_screening(index, data)
    with tab3:
        tab_data_status(data)
    with tab4:
        tab_about()


if __name__ == "__main__":
    main()
